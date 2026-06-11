from __future__ import annotations
import csv
import os
import itertools
from datetime import datetime
from typing import Any, Dict, List, Optional, Union, Tuple
from io import BytesIO, StringIO

from sqlalchemy import select, func, and_
from sqlalchemy.orm import Session
from fastapi import UploadFile, HTTPException, status

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.core.cache import cache
from app.models.import_export import (
    ImportJob,
    ImportError,
    ImportJobType,
    ImportStatus,
    FileType,
    ImportErrorCode,
)
from app.models.product import Product, ProductStatus
from app.models.sku import SKU, SkuStatus, SkuLifecycleStatus
from app.models.inventory import Inventory
from app.models.warehouse import Warehouse
from app.utils.helpers import generate_sku_code, safe_float, safe_int
from app.utils.exceptions import ValidationException, NotFoundException
from app.schemas.import_export import (
    ImportJobCreate,
    ImportJobUpdate,
    ImportErrorCreate,
    SkuExportFilter,
    ProductExportFilter,
)


EXCEL_COLUMNS = [
    "商品名称",
    "品类",
    "品牌",
    "SKU编码",
    "颜色",
    "尺寸",
    "成本价",
    "销售价",
    "状态",
    "安全库存",
    "最大库存",
]

ERROR_REPORT_COLUMNS = [
    "行号",
    "字段名",
    "错误码",
    "错误信息",
    "原始数据",
]


class ImportExportService:
    def __init__(self):
        self.cache_prefix = "import_job"
        self.error_report_dir = "app/data/error_reports"
        self.export_dir = "app/data/exports"
        os.makedirs(self.error_report_dir, exist_ok=True)
        os.makedirs(self.export_dir, exist_ok=True)

    def _read_excel(self, file_content: bytes) -> List[Dict[str, Any]]:
        workbook = load_workbook(filename=BytesIO(file_content), data_only=True)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        rows = []
        for row_idx in range(2, sheet.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers):
                row_data[str(header)] = sheet.cell(row=row_idx, column=col_idx + 1).value
            if any(row_data.values()):
                rows.append(row_data)
        return rows

    def _read_csv(self, file_content: bytes) -> List[Dict[str, Any]]:
        content = file_content.decode("utf-8-sig")
        reader = csv.DictReader(StringIO(content))
        return [dict(row) for row in reader]

    def _read_file(self, file: UploadFile, file_type: FileType) -> List[Dict[str, Any]]:
        content = file.file.read()
        if file_type == FileType.EXCEL:
            return self._read_excel(content)
        elif file_type == FileType.CSV:
            return self._read_csv(content)
        else:
            raise ValidationException(f"Unsupported file type: {file_type}")

    def _get_cell_value(self, row: Dict[str, Any], key: str) -> Optional[Any]:
        for k, v in row.items():
            if isinstance(k, str) and k.strip() == key:
                return v
        return row.get(key)

    def validate_row(
        self,
        db: Session,
        row: Dict[str, Any],
        row_number: int,
        job_type: ImportJobType,
        existing_sku_codes: set,
    ) -> Tuple[bool, List[ImportErrorCreate]]:
        errors = []

        product_name = self._get_cell_value(row, "商品名称")
        sku_code = self._get_cell_value(row, "SKU编码")
        cost_price = self._get_cell_value(row, "成本价")
        selling_price = self._get_cell_value(row, "销售价")
        status = self._get_cell_value(row, "状态")
        color = self._get_cell_value(row, "颜色")
        size = self._get_cell_value(row, "尺寸")

        if not product_name:
            errors.append(
                ImportErrorCreate(
                    job_id=0,
                    row_number=row_number,
                    field_name="商品名称",
                    error_code=ImportErrorCode.REQUIRED_FIELD_MISSING,
                    error_message="商品名称不能为空",
                    raw_data=row,
                )
            )

        if not cost_price:
            errors.append(
                ImportErrorCreate(
                    job_id=0,
                    row_number=row_number,
                    field_name="成本价",
                    error_code=ImportErrorCode.REQUIRED_FIELD_MISSING,
                    error_message="成本价不能为空",
                    raw_data=row,
                )
            )
        else:
            try:
                cost_price_val = safe_float(cost_price)
                if cost_price_val < 0:
                    errors.append(
                        ImportErrorCreate(
                            job_id=0,
                            row_number=row_number,
                            field_name="成本价",
                            error_code=ImportErrorCode.INVALID_PRICE,
                            error_message="成本价不能为负数",
                            raw_data=row,
                        )
                    )
            except (ValueError, TypeError):
                errors.append(
                    ImportErrorCreate(
                        job_id=0,
                        row_number=row_number,
                        field_name="成本价",
                        error_code=ImportErrorCode.INVALID_PRICE,
                        error_message="成本价格式不正确",
                        raw_data=row,
                    )
                )

        if selling_price is not None:
            try:
                selling_price_val = safe_float(selling_price)
                if selling_price_val < 0:
                    errors.append(
                        ImportErrorCreate(
                            job_id=0,
                            row_number=row_number,
                            field_name="销售价",
                            error_code=ImportErrorCode.INVALID_PRICE,
                            error_message="销售价不能为负数",
                            raw_data=row,
                        )
                    )
            except (ValueError, TypeError):
                errors.append(
                    ImportErrorCreate(
                        job_id=0,
                        row_number=row_number,
                        field_name="销售价",
                        error_code=ImportErrorCode.INVALID_PRICE,
                        error_message="销售价格式不正确",
                        raw_data=row,
                    )
                )

        if status:
            valid_statuses = [s.value for s in ProductStatus] + [s.value for s in SkuStatus]
            if str(status).upper() not in valid_statuses:
                errors.append(
                    ImportErrorCreate(
                        job_id=0,
                        row_number=row_number,
                        field_name="状态",
                        error_code=ImportErrorCode.INVALID_STATUS,
                        error_message=f"状态值无效，有效值为: {', '.join(valid_statuses)}",
                        raw_data=row,
                    )
                )

        if sku_code and job_type == ImportJobType.SKU_IMPORT:
            if str(sku_code) in existing_sku_codes:
                errors.append(
                    ImportErrorCreate(
                        job_id=0,
                        row_number=row_number,
                        field_name="SKU编码",
                        error_code=ImportErrorCode.DUPLICATE_SKU_CODE,
                        error_message=f"SKU编码 {sku_code} 在文件中重复",
                        raw_data=row,
                    )
                )
            else:
                existing_sku = db.execute(
                    select(SKU).where(SKU.sku_code == str(sku_code))
                ).scalar_one_or_none()
                if existing_sku:
                    errors.append(
                        ImportErrorCreate(
                            job_id=0,
                            row_number=row_number,
                            field_name="SKU编码",
                            error_code=ImportErrorCode.DUPLICATE_SKU_CODE,
                            error_message=f"SKU编码 {sku_code} 已存在",
                            raw_data=row,
                        )
                    )

        if color or size:
            if not str(color or "") and not str(size or ""):
                errors.append(
                    ImportErrorCreate(
                        job_id=0,
                        row_number=row_number,
                        field_name="属性",
                        error_code=ImportErrorCode.INVALID_ATTRIBUTE,
                        error_message="颜色和尺寸不能同时为空",
                        raw_data=row,
                    )
                )

        return len(errors) == 0, errors

    def _generate_sku_codes_from_attributes(
        self,
        product_id: int,
        row: Dict[str, Any],
    ) -> List[Tuple[str, Dict[str, Any]]]:
        color = self._get_cell_value(row, "颜色")
        size = self._get_cell_value(row, "尺寸")

        color_values = []
        size_values = []

        if color:
            color_values = [c.strip() for c in str(color).split(",") if c.strip()]
        if size:
            size_values = [s.strip() for s in str(size).split(",") if s.strip()]

        if not color_values and not size_values:
            return []

        combinations = []
        if color_values and size_values:
            for c in color_values:
                for s in size_values:
                    attributes = {}
                    if c:
                        attributes["color"] = {"value": c, "name": "颜色"}
                    if s:
                        attributes["size"] = {"value": s, "name": "尺寸"}
                    sku_code = generate_sku_code(product_id, attributes)
                    combinations.append((sku_code, attributes))
        elif color_values:
            for c in color_values:
                attributes = {"color": {"value": c, "name": "颜色"}}
                sku_code = generate_sku_code(product_id, attributes)
                combinations.append((sku_code, attributes))
        elif size_values:
            for s in size_values:
                attributes = {"size": {"value": s, "name": "尺寸"}}
                sku_code = generate_sku_code(product_id, attributes)
                combinations.append((sku_code, attributes))

        return combinations

    def create_job(
        self,
        db: Session,
        *,
        obj_in: ImportJobCreate,
    ) -> ImportJob:
        db_job = ImportJob(**obj_in.model_dump())
        db.add(db_job)
        db.flush()
        db.refresh(db_job)
        return db_job

    def update_job(
        self,
        db: Session,
        *,
        db_obj: ImportJob,
        obj_in: ImportJobUpdate,
    ) -> ImportJob:
        update_data = obj_in.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(db_obj, field, value)
        db.flush()
        db.refresh(db_obj)
        cache.delete(f"{self.cache_prefix}:{db_obj.id}")
        return db_obj

    def get_job(
        self,
        db: Session,
        *,
        job_id: int,
    ) -> ImportJob:
        job = db.execute(
            select(ImportJob).where(ImportJob.id == job_id)
        ).scalar_one_or_none()
        if not job:
            raise NotFoundException("ImportJob", job_id)
        return job

    def generate_error_report(
        self,
        db: Session,
        *,
        job_id: int,
        errors: List[ImportError],
    ) -> str:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "错误报告"

        header_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        header_font = Font(bold=True, color="9C0006")

        for col_idx, header in enumerate(ERROR_REPORT_COLUMNS, 1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, error in enumerate(errors, 2):
            sheet.cell(row=row_idx, column=1, value=error.row_number)
            sheet.cell(row=row_idx, column=2, value=error.field_name)
            sheet.cell(row=row_idx, column=3, value=error.error_code.value)
            sheet.cell(row=row_idx, column=4, value=error.error_message)
            raw_data_str = str(error.raw_data) if error.raw_data else ""
            sheet.cell(row=row_idx, column=5, value=raw_data_str)

        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column].width = adjusted_width

        file_name = f"error_report_job_{job_id}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(self.error_report_dir, file_name)
        workbook.save(file_path)

        job = self.get_job(db, job_id=job_id)
        job.error_report_path = file_path
        db.flush()

        return file_path

    def import_products_from_file(
        self,
        db: Session,
        *,
        file: UploadFile,
        file_type: FileType,
        created_by: Optional[int] = None,
    ) -> Dict[str, Any]:
        job = self.create_job(
            db,
            obj_in=ImportJobCreate(
                job_type=ImportJobType.PRODUCT_IMPORT,
                file_name=file.filename or "unknown",
                file_type=file_type,
                created_by=created_by,
            ),
        )

        try:
            self.update_job(
                db,
                db_obj=job,
                obj_in=ImportJobUpdate(status=ImportStatus.PROCESSING),
            )

            rows = self._read_file(file, file_type)
            total_count = len(rows)

            self.update_job(
                db,
                db_obj=job,
                obj_in=ImportJobUpdate(total_count=total_count),
            )

            created_products: List[Product] = []
            created_skus: List[SKU] = []
            all_errors: List[ImportError] = []
            existing_sku_codes: set = set()

            for row_idx, row in enumerate(rows, start=2):
                is_valid, row_errors = self.validate_row(
                    db, row, row_idx, ImportJobType.PRODUCT_IMPORT, existing_sku_codes
                )

                if not is_valid:
                    for err in row_errors:
                        err.job_id = job.id
                        db_error = ImportError(**err.model_dump())
                        db.add(db_error)
                        all_errors.append(db_error)
                    continue

                try:
                    product_name = str(self._get_cell_value(row, "商品名称"))
                    category = self._get_cell_value(row, "品类")
                    brand = self._get_cell_value(row, "品牌")
                    sku_code = self._get_cell_value(row, "SKU编码")
                    cost_price = safe_float(self._get_cell_value(row, "成本价"))
                    selling_price = safe_float(self._get_cell_value(row, "销售价"))
                    status_str = self._get_cell_value(row, "状态")
                    safety_stock = safe_int(self._get_cell_value(row, "安全库存"))
                    maximum_stock = safe_int(self._get_cell_value(row, "最大库存"))

                    product_status = ProductStatus.DRAFT
                    if status_str:
                        status_upper = str(status_str).upper()
                        if status_upper in [s.value for s in ProductStatus]:
                            product_status = ProductStatus(status_upper)

                    product = Product(
                        name=product_name,
                        category=str(category) if category else None,
                        brand=str(brand) if brand else None,
                        status=product_status,
                    )
                    db.add(product)
                    db.flush()
                    db.refresh(product)
                    created_products.append(product)

                    if sku_code:
                        sku_status = SkuStatus.DRAFT
                        if status_str:
                            status_upper = str(status_str).upper()
                            if status_upper in [s.value for s in SkuStatus]:
                                sku_status = SkuStatus(status_upper)

                        attributes = {}
                        color = self._get_cell_value(row, "颜色")
                        size = self._get_cell_value(row, "尺寸")
                        if color:
                            attributes["color"] = {"value": str(color), "name": "颜色"}
                        if size:
                            attributes["size"] = {"value": str(size), "name": "尺寸"}

                        sku = SKU(
                            sku_code=str(sku_code),
                            product_id=product.id,
                            attributes=attributes if attributes else None,
                            cost_price=cost_price,
                            selling_price=selling_price,
                            status=sku_status,
                            safety_stock=safety_stock,
                            maximum_stock=maximum_stock,
                        )
                        db.add(sku)
                        db.flush()
                        db.refresh(sku)
                        created_skus.append(sku)
                        existing_sku_codes.add(str(sku_code))
                    else:
                        sku_combinations = self._generate_sku_codes_from_attributes(product.id, row)
                        for generated_sku_code, attributes in sku_combinations:
                            if generated_sku_code in existing_sku_codes:
                                continue
                            existing_sku = db.execute(
                                select(SKU).where(SKU.sku_code == generated_sku_code)
                            ).scalar_one_or_none()
                            if existing_sku:
                                continue

                            sku_status = SkuStatus.DRAFT
                            if status_str:
                                status_upper = str(status_str).upper()
                                if status_upper in [s.value for s in SkuStatus]:
                                    sku_status = SkuStatus(status_upper)

                            sku = SKU(
                                sku_code=generated_sku_code,
                                product_id=product.id,
                                attributes=attributes,
                                cost_price=cost_price,
                                selling_price=selling_price,
                                status=sku_status,
                                safety_stock=safety_stock,
                                maximum_stock=maximum_stock,
                            )
                            db.add(sku)
                            db.flush()
                            db.refresh(sku)
                            created_skus.append(sku)
                            existing_sku_codes.add(generated_sku_code)

                except Exception as e:
                    db_error = ImportError(
                        job_id=job.id,
                        row_number=row_idx,
                        field_name=None,
                        error_code=ImportErrorCode.INVALID_ATTRIBUTE,
                        error_message=f"处理行时发生错误: {str(e)}",
                        raw_data=row,
                    )
                    db.add(db_error)
                    all_errors.append(db_error)

            success_count = len(created_products)
            failed_count = len(all_errors)

            self.update_job(
                db,
                db_obj=job,
                obj_in=ImportJobUpdate(
                    status=ImportStatus.COMPLETED,
                    success_count=success_count,
                    failed_count=failed_count,
                    completed_at=datetime.utcnow(),
                ),
            )

            if all_errors:
                self.generate_error_report(db, job_id=job.id, errors=all_errors)

            db.commit()

            cache.delete_pattern("product:list:*")
            cache.delete_pattern("sku:list:*")

            return {
                "job_id": job.id,
                "total_count": total_count,
                "success_count": success_count,
                "failed_count": failed_count,
                "status": ImportStatus.COMPLETED,
                "created_products": created_products,
                "created_skus": created_skus,
                "errors": all_errors,
            }

        except Exception as e:
            db.rollback()
            self.update_job(
                db,
                db_obj=job,
                obj_in=ImportJobUpdate(
                    status=ImportStatus.FAILED,
                    completed_at=datetime.utcnow(),
                ),
            )
            db.commit()
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Import failed: {str(e)}",
            )

    def import_skus_from_file(
        self,
        db: Session,
        *,
        file: UploadFile,
        file_type: FileType,
        created_by: Optional[int] = None,
    ) -> Dict[str, Any]:
        job = self.create_job(
            db,
            obj_in=ImportJobCreate(
                job_type=ImportJobType.SKU_IMPORT,
                file_name=file.filename or "unknown",
                file_type=file_type,
                created_by=created_by,
            ),
        )

        try:
            self.update_job(
                db,
                db_obj=job,
                obj_in=ImportJobUpdate(status=ImportStatus.PROCESSING),
            )

            rows = self._read_file(file, file_type)
            total_count = len(rows)

            self.update_job(
                db,
                db_obj=job,
                obj_in=ImportJobUpdate(total_count=total_count),
            )

            created_products: List[Product] = []
            created_skus: List[SKU] = []
            all_errors: List[ImportError] = []
            existing_sku_codes: set = set()

            for row_idx, row in enumerate(rows, start=2):
                is_valid, row_errors = self.validate_row(
                    db, row, row_idx, ImportJobType.SKU_IMPORT, existing_sku_codes
                )

                if not is_valid:
                    for err in row_errors:
                        err.job_id = job.id
                        db_error = ImportError(**err.model_dump())
                        db.add(db_error)
                        all_errors.append(db_error)
                    continue

                try:
                    product_name = str(self._get_cell_value(row, "商品名称"))
                    category = self._get_cell_value(row, "品类")
                    brand = self._get_cell_value(row, "品牌")
                    sku_code = self._get_cell_value(row, "SKU编码")
                    cost_price = safe_float(self._get_cell_value(row, "成本价"))
                    selling_price = safe_float(self._get_cell_value(row, "销售价"))
                    status_str = self._get_cell_value(row, "状态")
                    safety_stock = safe_int(self._get_cell_value(row, "安全库存"))
                    maximum_stock = safe_int(self._get_cell_value(row, "最大库存"))

                    product = db.execute(
                        select(Product).where(Product.name == product_name)
                    ).scalar_one_or_none()

                    if not product:
                        product_status = ProductStatus.DRAFT
                        if status_str:
                            status_upper = str(status_str).upper()
                            if status_upper in [s.value for s in ProductStatus]:
                                product_status = ProductStatus(status_upper)

                        product = Product(
                            name=product_name,
                            category=str(category) if category else None,
                            brand=str(brand) if brand else None,
                            status=product_status,
                        )
                        db.add(product)
                        db.flush()
                        db.refresh(product)
                        created_products.append(product)

                    if sku_code:
                        sku_status = SkuStatus.DRAFT
                        if status_str:
                            status_upper = str(status_str).upper()
                            if status_upper in [s.value for s in SkuStatus]:
                                sku_status = SkuStatus(status_upper)

                        attributes = {}
                        color = self._get_cell_value(row, "颜色")
                        size = self._get_cell_value(row, "尺寸")
                        if color:
                            attributes["color"] = {"value": str(color), "name": "颜色"}
                        if size:
                            attributes["size"] = {"value": str(size), "name": "尺寸"}

                        sku = SKU(
                            sku_code=str(sku_code),
                            product_id=product.id,
                            attributes=attributes if attributes else None,
                            cost_price=cost_price,
                            selling_price=selling_price,
                            status=sku_status,
                            safety_stock=safety_stock,
                            maximum_stock=maximum_stock,
                        )
                        db.add(sku)
                        db.flush()
                        db.refresh(sku)
                        created_skus.append(sku)
                        existing_sku_codes.add(str(sku_code))
                    else:
                        sku_combinations = self._generate_sku_codes_from_attributes(product.id, row)
                        for generated_sku_code, attributes in sku_combinations:
                            if generated_sku_code in existing_sku_codes:
                                continue
                            existing_sku = db.execute(
                                select(SKU).where(SKU.sku_code == generated_sku_code)
                            ).scalar_one_or_none()
                            if existing_sku:
                                continue

                            sku_status = SkuStatus.DRAFT
                            if status_str:
                                status_upper = str(status_str).upper()
                                if status_upper in [s.value for s in SkuStatus]:
                                    sku_status = SkuStatus(status_upper)

                            sku = SKU(
                                sku_code=generated_sku_code,
                                product_id=product.id,
                                attributes=attributes,
                                cost_price=cost_price,
                                selling_price=selling_price,
                                status=sku_status,
                                safety_stock=safety_stock,
                                maximum_stock=maximum_stock,
                            )
                            db.add(sku)
                            db.flush()
                            db.refresh(sku)
                            created_skus.append(sku)
                            existing_sku_codes.add(generated_sku_code)

                except Exception as e:
                    db_error = ImportError(
                        job_id=job.id,
                        row_number=row_idx,
                        field_name=None,
                        error_code=ImportErrorCode.INVALID_ATTRIBUTE,
                        error_message=f"处理行时发生错误: {str(e)}",
                        raw_data=row,
                    )
                    db.add(db_error)
                    all_errors.append(db_error)

            success_count = len(created_skus)
            failed_count = len(all_errors)

            self.update_job(
                db,
                db_obj=job,
                obj_in=ImportJobUpdate(
                    status=ImportStatus.COMPLETED,
                    success_count=success_count,
                    failed_count=failed_count,
                    completed_at=datetime.utcnow(),
                ),
            )

            if all_errors:
                self.generate_error_report(db, job_id=job.id, errors=all_errors)

            db.commit()

            cache.delete_pattern("product:list:*")
            cache.delete_pattern("sku:list:*")

            return {
                "job_id": job.id,
                "total_count": total_count,
                "success_count": success_count,
                "failed_count": failed_count,
                "status": ImportStatus.COMPLETED,
                "created_products": created_products,
                "created_skus": created_skus,
                "errors": all_errors,
            }

        except Exception as e:
            db.rollback()
            self.update_job(
                db,
                db_obj=job,
                obj_in=ImportJobUpdate(
                    status=ImportStatus.FAILED,
                    completed_at=datetime.utcnow(),
                ),
            )
            db.commit()
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Import failed: {str(e)}",
            )

    def export_skus(
        self,
        db: Session,
        *,
        filters: SkuExportFilter,
    ) -> Tuple[str, int]:
        stmt = (
            select(
                SKU,
                Product.name.label("product_name"),
                Product.category.label("product_category"),
                Product.brand.label("product_brand"),
                func.coalesce(func.sum(Inventory.quantity), 0).label("current_stock"),
                func.coalesce(func.sum(Inventory.available_quantity), 0).label("available_stock"),
                Warehouse.name.label("warehouse_name"),
            )
            .join(Product, SKU.product_id == Product.id)
            .outerjoin(Inventory, SKU.id == Inventory.sku_id)
            .outerjoin(Warehouse, Inventory.warehouse_id == Warehouse.id)
        )

        where_conditions = []
        if filters.category:
            where_conditions.append(Product.category == filters.category)
        if filters.warehouse_id:
            where_conditions.append(Inventory.warehouse_id == filters.warehouse_id)
        if filters.product_id:
            where_conditions.append(SKU.product_id == filters.product_id)
        if filters.sku_code:
            where_conditions.append(SKU.sku_code.like(f"%{filters.sku_code}%"))
        if filters.status:
            where_conditions.append(SKU.status == filters.status)

        if filters.stock_status:
            stock_status = filters.stock_status.upper()
            if stock_status == "IN_STOCK":
                where_conditions.append(func.coalesce(func.sum(Inventory.quantity), 0) > 0)
            elif stock_status == "LOW_STOCK":
                where_conditions.append(
                    and_(
                        func.coalesce(func.sum(Inventory.quantity), 0) > 0,
                        func.coalesce(func.sum(Inventory.quantity), 0) <= SKU.safety_stock,
                    )
                )
            elif stock_status == "OUT_OF_STOCK":
                where_conditions.append(func.coalesce(func.sum(Inventory.quantity), 0) == 0)

        if where_conditions:
            stmt = stmt.where(and_(*where_conditions))

        stmt = stmt.group_by(
            SKU.id,
            Product.name,
            Product.category,
            Product.brand,
            Warehouse.name,
        )

        results = db.execute(stmt).all()

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SKU导出"

        export_columns = [
            "SKU编码",
            "商品名称",
            "品类",
            "品牌",
            "颜色",
            "尺寸",
            "成本价",
            "销售价",
            "状态",
            "安全库存",
            "最大库存",
            "当前库存",
            "可用库存",
            "仓库名称",
        ]

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, header in enumerate(export_columns, 1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, result in enumerate(results, 2):
            sku = result.SKU
            color = ""
            size = ""
            if sku.attributes:
                if isinstance(sku.attributes, dict):
                    if "color" in sku.attributes:
                        color = sku.attributes["color"].get("value", "") if isinstance(sku.attributes["color"], dict) else str(sku.attributes["color"])
                    if "size" in sku.attributes:
                        size = sku.attributes["size"].get("value", "") if isinstance(sku.attributes["size"], dict) else str(sku.attributes["size"])

            sheet.cell(row=row_idx, column=1, value=sku.sku_code)
            sheet.cell(row=row_idx, column=2, value=result.product_name)
            sheet.cell(row=row_idx, column=3, value=result.product_category or "")
            sheet.cell(row=row_idx, column=4, value=result.product_brand or "")
            sheet.cell(row=row_idx, column=5, value=color)
            sheet.cell(row=row_idx, column=6, value=size)
            sheet.cell(row=row_idx, column=7, value=sku.cost_price)
            sheet.cell(row=row_idx, column=8, value=sku.selling_price)
            sheet.cell(row=row_idx, column=9, value=sku.status.value)
            sheet.cell(row=row_idx, column=10, value=sku.safety_stock)
            sheet.cell(row=row_idx, column=11, value=sku.maximum_stock)
            sheet.cell(row=row_idx, column=12, value=result.current_stock)
            sheet.cell(row=row_idx, column=13, value=result.available_stock)
            sheet.cell(row=row_idx, column=14, value=result.warehouse_name or "")

        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column].width = adjusted_width

        file_name = f"skus_export_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(self.export_dir, file_name)
        workbook.save(file_path)

        return file_path, len(results)

    def export_products(
        self,
        db: Session,
        *,
        filters: ProductExportFilter,
    ) -> Tuple[str, int]:
        stmt = select(Product).order_by(Product.id)

        where_conditions = []
        if filters.category:
            where_conditions.append(Product.category == filters.category)
        if filters.brand:
            where_conditions.append(Product.brand == filters.brand)
        if filters.status:
            where_conditions.append(Product.status == filters.status)

        if where_conditions:
            stmt = stmt.where(and_(*where_conditions))

        products = db.execute(stmt).scalars().all()

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "商品导出"

        export_columns = [
            "商品ID",
            "商品名称",
            "品类",
            "品牌",
            "状态",
            "SKU编码",
            "颜色",
            "尺寸",
            "成本价",
            "销售价",
            "SKU状态",
            "安全库存",
            "最大库存",
        ]

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, header in enumerate(export_columns, 1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row_idx = 2
        total_count = 0

        for product in products:
            if filters.include_skus and product.skus:
                for sku in product.skus:
                    color = ""
                    size = ""
                    if sku.attributes:
                        if isinstance(sku.attributes, dict):
                            if "color" in sku.attributes:
                                color = sku.attributes["color"].get("value", "") if isinstance(sku.attributes["color"], dict) else str(sku.attributes["color"])
                            if "size" in sku.attributes:
                                size = sku.attributes["size"].get("value", "") if isinstance(sku.attributes["size"], dict) else str(sku.attributes["size"])

                    sheet.cell(row=row_idx, column=1, value=product.id)
                    sheet.cell(row=row_idx, column=2, value=product.name)
                    sheet.cell(row=row_idx, column=3, value=product.category or "")
                    sheet.cell(row=row_idx, column=4, value=product.brand or "")
                    sheet.cell(row=row_idx, column=5, value=product.status.value)
                    sheet.cell(row=row_idx, column=6, value=sku.sku_code)
                    sheet.cell(row=row_idx, column=7, value=color)
                    sheet.cell(row=row_idx, column=8, value=size)
                    sheet.cell(row=row_idx, column=9, value=sku.cost_price)
                    sheet.cell(row=row_idx, column=10, value=sku.selling_price)
                    sheet.cell(row=row_idx, column=11, value=sku.status.value)
                    sheet.cell(row=row_idx, column=12, value=sku.safety_stock)
                    sheet.cell(row=row_idx, column=13, value=sku.maximum_stock)
                    row_idx += 1
                    total_count += 1
            else:
                sheet.cell(row=row_idx, column=1, value=product.id)
                sheet.cell(row=row_idx, column=2, value=product.name)
                sheet.cell(row=row_idx, column=3, value=product.category or "")
                sheet.cell(row=row_idx, column=4, value=product.brand or "")
                sheet.cell(row=row_idx, column=5, value=product.status.value)
                row_idx += 1
                total_count += 1

        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column].width = adjusted_width

        file_name = f"products_export_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(self.export_dir, file_name)
        workbook.save(file_path)

        return file_path, total_count


import_export_service = ImportExportService()
