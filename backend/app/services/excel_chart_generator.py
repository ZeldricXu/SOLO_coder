import os
import tempfile
import pandas as pd
import numpy as np
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass
from enum import Enum
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart
    from openpyxl.chart.reference import Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

class ChartType(str, Enum):
    BAR = "bar"
    PIE = "pie"
    LINE = "line"
    SCATTER = "scatter"
    STACKED_BAR = "stacked_bar"
    HORIZONTAL_BAR = "horizontal_bar"

@dataclass
class ChartConfig:
    chart_type: ChartType
    title: str
    x_axis_title: str
    y_axis_title: str
    data_series: List[Dict[str, Any]]
    categories: List[str]
    width: int = 15
    height: int = 10
    colors: List[str] = None
    show_legend: bool = True
    show_data_labels: bool = False
    
    def __post_init__(self):
        if self.colors is None:
            self.colors = [
                '4472C4', 'ED7D31', '70AD47', '7030A0',
                'FFC000', '5B9BD5', 'C00000', '00B050'
            ]

class ExcelChartGenerator:
    """
    Excel图表生成器
    用于创建可编辑的Excel图表，支持作为OLE对象嵌入Word
    """
    
    def __init__(self, temp_folder: str = None):
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required for Excel chart generation")
        
        self.temp_folder = temp_folder or tempfile.gettempdir()
        os.makedirs(self.temp_folder, exist_ok=True)
    
    def _create_worksheet_with_data(
        self, 
        wb: Workbook, 
        sheet_name: str, 
        config: ChartConfig
    ) -> Tuple['Worksheet', Tuple[int, int]]:
        """
        在工作簿中创建工作表并写入数据
        
        Returns:
            (worksheet, (data_start_row, data_end_col))
        """
        ws = wb.create_sheet(sheet_name)
        
        ws.cell(row=1, column=1, value=config.title)
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(config.categories) + 1)
        
        header_row = 3
        ws.cell(row=header_row, column=1, value="类别")
        ws.cell(row=header_row, column=1).font = Font(bold=True)
        
        for col_idx, category in enumerate(config.categories, start=2):
            cell = ws.cell(row=header_row, column=col_idx, value=category)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        data_start_row = header_row + 1
        for series_idx, series in enumerate(config.data_series):
            row_idx = data_start_row + series_idx
            ws.cell(row=row_idx, column=1, value=series.get('name', f'系列{series_idx + 1}'))
            ws.cell(row=row_idx, column=1).font = Font(bold=True)
            
            for col_idx, value in enumerate(series.get('values', []), start=2):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='center')
        
        for col_idx in range(1, len(config.categories) + 2):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 15
        
        data_end_row = data_start_row + len(config.data_series) - 1
        data_end_col = len(config.categories) + 1
        
        return ws, (data_start_row, data_end_col)
    
    def _create_bar_chart(
        self, 
        ws: 'Worksheet',
        data_range: Tuple[int, int],
        config: ChartConfig,
        chart_type: str = "bar"
    ) -> 'BarChart':
        """创建柱状图或堆叠柱状图"""
        data_start_row, data_end_col = data_range
        
        if chart_type == "stacked_bar":
            chart = BarChart()
            chart.type = "col"
            chart.grouping = "stacked"
        elif chart_type == "horizontal_bar":
            chart = BarChart()
            chart.type = "bar"
        else:
            chart = BarChart()
            chart.type = "col"
        
        chart.title = config.title
        chart.x_axis.title = config.x_axis_title
        chart.y_axis.title = config.y_axis_title
        
        data_end_row = data_start_row + len(config.data_series)
        data = Reference(ws, min_col=2, min_row=data_start_row, max_col=data_end_col, max_row=data_end_row)
        categories = Reference(ws, min_col=2, min_row=data_start_row - 1, max_col=data_end_col, max_row=data_start_row - 1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        
        chart.shape = 4
        chart.style = 10
        
        for idx, series in enumerate(chart.series):
            if idx < len(config.colors):
                fill = PatternFill(start_color=config.colors[idx], end_color=config.colors[idx], fill_type="solid")
                series.graphicalProperties.solidFill = config.colors[idx]
        
        chart.width = config.width
        chart.height = config.height
        
        return chart
    
    def _create_pie_chart(
        self, 
        ws: 'Worksheet',
        data_range: Tuple[int, int],
        config: ChartConfig
    ) -> 'PieChart':
        """创建饼图"""
        data_start_row, data_end_col = data_range
        
        chart = PieChart()
        chart.title = config.title
        
        data_end_row = data_start_row + len(config.data_series)
        data = Reference(ws, min_col=2, min_row=data_start_row, max_col=data_end_col, max_row=data_end_row)
        categories = Reference(ws, min_col=2, min_row=data_start_row - 1, max_col=data_end_col, max_row=data_start_row - 1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        
        chart.style = 26
        
        for idx, point in enumerate(chart.series[0].data_points):
            if idx < len(config.colors):
                dp = DataPoint(idx=idx)
                dp.graphicalProperties.solidFill = config.colors[idx]
                chart.series[0].data_points.append(dp)
        
        chart.width = config.width
        chart.height = config.height
        
        return chart
    
    def _create_line_chart(
        self, 
        ws: 'Worksheet',
        data_range: Tuple[int, int],
        config: ChartConfig
    ) -> 'LineChart':
        """创建折线图"""
        data_start_row, data_end_col = data_range
        
        chart = LineChart()
        chart.title = config.title
        chart.x_axis.title = config.x_axis_title
        chart.y_axis.title = config.y_axis_title
        
        data_end_row = data_start_row + len(config.data_series)
        data = Reference(ws, min_col=2, min_row=data_start_row, max_col=data_end_col, max_row=data_end_row)
        categories = Reference(ws, min_col=2, min_row=data_start_row - 1, max_col=data_end_col, max_row=data_start_row - 1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        
        chart.style = 10
        chart.width = config.width
        chart.height = config.height
        
        return chart
    
    def generate_excel_chart(
        self,
        config: ChartConfig,
        filename: str = None
    ) -> str:
        """
        生成包含图表的Excel文件
        
        Args:
            config: 图表配置
            filename: 输出文件名（可选）
            
        Returns:
            生成的Excel文件路径
        """
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"chart_{timestamp}.xlsx"
        
        filepath = os.path.join(self.temp_folder, filename)
        
        wb = Workbook()
        
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        ws, data_range = self._create_worksheet_with_data(wb, "图表数据", config)
        
        if config.chart_type == ChartType.BAR:
            chart = self._create_bar_chart(ws, data_range, config, "bar")
        elif config.chart_type == ChartType.STACKED_BAR:
            chart = self._create_bar_chart(ws, data_range, config, "stacked_bar")
        elif config.chart_type == ChartType.HORIZONTAL_BAR:
            chart = self._create_bar_chart(ws, data_range, config, "horizontal_bar")
        elif config.chart_type == ChartType.PIE:
            chart = self._create_pie_chart(ws, data_range, config)
        elif config.chart_type == ChartType.LINE:
            chart = self._create_line_chart(ws, data_range, config)
        else:
            chart = self._create_bar_chart(ws, data_range, config, "bar")
        
        ws.add_chart(chart, "A10")
        
        summary_sheet = wb.create_sheet("图表摘要")
        summary_sheet.cell(row=1, column=1, value="图表生成摘要")
        summary_sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
        
        summary_data = [
            ("图表标题", config.title),
            ("图表类型", config.chart_type.value),
            ("X轴标题", config.x_axis_title),
            ("Y轴标题", config.y_axis_title),
            ("数据系列数", len(config.data_series)),
            ("类别数", len(config.categories)),
            ("生成时间", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        ]
        
        for idx, (key, value) in enumerate(summary_data, start=3):
            summary_sheet.cell(row=idx, column=1, value=key)
            summary_sheet.cell(row=idx, column=1).font = Font(bold=True)
            summary_sheet.cell(row=idx, column=2, value=str(value))
        
        summary_sheet.column_dimensions['A'].width = 20
        summary_sheet.column_dimensions['B'].width = 40
        
        wb.save(filepath)
        
        return filepath
    
    def create_chart_from_analysis_result(
        self,
        analysis_type: str,
        analysis_data: Dict[str, Any],
        chart_type: ChartType = None
    ) -> ChartConfig:
        """
        从分析结果创建图表配置
        
        Args:
            analysis_type: 分析类型（frequency, descriptive, cross）
            analysis_data: 分析结果数据
            chart_type: 图表类型（可选）
            
        Returns:
            ChartConfig对象
        """
        if analysis_type == "frequency":
            return self._create_frequency_chart_config(analysis_data, chart_type)
        elif analysis_type == "cross":
            return self._create_cross_chart_config(analysis_data, chart_type)
        elif analysis_type == "descriptive":
            return self._create_descriptive_chart_config(analysis_data, chart_type)
        else:
            return self._create_default_chart_config(analysis_data)
    
    def _create_frequency_chart_config(
        self,
        data: Dict[str, Any],
        chart_type: ChartType = None
    ) -> ChartConfig:
        """从频数分析结果创建图表配置"""
        frequencies = data.get("frequencies", [])
        field_name = data.get("field_name", "频数分布")
        
        categories = [str(f.get("value", "")) for f in frequencies]
        values = [f.get("count", 0) for f in frequencies]
        
        if chart_type is None:
            if len(categories) <= 5:
                chart_type = ChartType.PIE
            else:
                chart_type = ChartType.BAR
        
        return ChartConfig(
            chart_type=chart_type,
            title=f"{field_name} - 频数分布",
            x_axis_title="选项",
            y_axis_title="数量",
            categories=categories[:15],
            data_series=[{"name": "数量", "values": values[:15]}]
        )
    
    def _create_cross_chart_config(
        self,
        data: Dict[str, Any],
        chart_type: ChartType = None
    ) -> ChartConfig:
        """从交叉分析结果创建图表配置"""
        cross_table = data.get("cross_table", [])
        variables = data.get("variables", ["变量1", "变量2"])
        
        if chart_type is None:
            chart_type = ChartType.STACKED_BAR
        
        categories = []
        data_series = []
        
        if cross_table:
            first_row = cross_table[0]
            col_values = first_row.get("col_values", {})
            categories = [k for k in col_values.keys() if k != "_total"]
            
            for row in cross_table:
                row_name = row.get("row", "")
                values = []
                for cat in categories:
                    val = col_values.get(cat, {})
                    if "count" in val:
                        values.append(val.get("count", 0))
                    elif "mean" in val:
                        values.append(val.get("mean", 0))
                    else:
                        values.append(0)
                
                data_series.append({"name": row_name, "values": values})
        
        return ChartConfig(
            chart_type=chart_type,
            title=f"交叉分析 - {variables[0]} vs {variables[1]}",
            x_axis_title=variables[1] if len(variables) > 1 else "列变量",
            y_axis_title="数值",
            categories=categories,
            data_series=data_series
        )
    
    def _create_descriptive_chart_config(
        self,
        data: Dict[str, Any],
        chart_type: ChartType = None
    ) -> ChartConfig:
        """从描述性统计结果创建图表配置"""
        field_name = data.get("field_name", "描述性统计")
        
        stats_keys = ["mean", "median", "std", "min", "max", "q25", "q75"]
        stats_labels = ["均值", "中位数", "标准差", "最小值", "最大值", "25分位", "75分位"]
        
        categories = stats_labels
        values = [data.get(key, 0) for key in stats_keys]
        
        if chart_type is None:
            chart_type = ChartType.BAR
        
        return ChartConfig(
            chart_type=chart_type,
            title=f"{field_name} - 描述性统计",
            x_axis_title="统计指标",
            y_axis_title="数值",
            categories=categories,
            data_series=[{"name": field_name, "values": values}]
        )
    
    def _create_default_chart_config(self, data: Dict[str, Any]) -> ChartConfig:
        """创建默认图表配置"""
        return ChartConfig(
            chart_type=ChartType.BAR,
            title="数据分析图表",
            x_axis_title="类别",
            y_axis_title="数值",
            categories=["类别1", "类别2", "类别3"],
            data_series=[{"name": "系列1", "values": [10, 20, 15]}]
        )
