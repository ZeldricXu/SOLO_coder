import io
import csv
import uuid
import time
from typing import Optional, Dict, Any, List, Callable
from datetime import datetime
from sqlalchemy.orm import Session
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reporthub.models import Report, ExportConfig
from reporthub.modules.storage_module import StorageModule
from reporthub.modules.statistics_module import StatisticsModule
from reporthub.modules.task_queue import TaskQueue, TaskStatus
from reporthub.modules.redis_module import (
    RedisExportQueue,
    is_redis_available
)
from reporthub.config.settings import settings


class ConfigurableExportModule:
    SUPPORTED_FORMATS = ["xlsx", "pdf", "csv"]

    def __init__(self, db: Session, storage_module: StorageModule, statistics_module: StatisticsModule):
        self.db = db
        self.storage_module = storage_module
        self.statistics_module = statistics_module
        self._format_handlers: Dict[str, Callable] = {
            "xlsx": self._export_xlsx,
            "pdf": self._export_pdf,
            "csv": self._export_csv
        }

    def get_supported_formats(self) -> List[str]:
        return settings.get_supported_export_formats()

    def is_format_supported(self, format_name: str) -> bool:
        return settings.get_export_format_config(format_name.lower()) is not None

    def get_format_config(self, format_name: str) -> Optional[Dict[str, Any]]:
        config = settings.get_export_format_config(format_name.lower())
        return config.to_dict() if config else None

    def export_report(self, report: Report, export_format: str,
                      export_options: Optional[Dict[str, Any]] = None) -> str:
        format_name = export_format.lower()
        if not self.is_format_supported(format_name):
            raise ValueError(f"不支持的导出格式: {export_format}")
        format_config = settings.get_export_format_config(format_name)
        base_options = format_config.options.copy() if format_config else {}
        if export_options:
            format_specific = export_options.get(format_name, {})
            base_options.update(format_specific)
        merged_options = {format_name: base_options}
        handler = self._format_handlers.get(format_name)
        if not handler:
            raise ValueError(f"没有为格式 {format_name} 注册导出处理器")
        file_data = handler(report, base_options)
        file_extension = format_config.file_extension if format_config else format_name
        file_name = f"{report.report_id}.{file_extension}"
        file_path = self.storage_module.save_export_file(file_name, file_data)
        self.statistics_module.update_export_stats(report.template_id)
        return file_path

    def batch_export(self, reports, export_format: str,
                     export_options: Optional[Dict[str, Any]] = None) -> Dict[str, str]:
        results = {}
        for report in reports:
            try:
                file_path = self.export_report(report, export_format, export_options)
                results[report.report_id] = file_path
            except Exception as e:
                results[report.report_id] = f"导出失败: {str(e)}"
        return results

    def _export_xlsx(self, report: Report, options: Dict[str, Any]) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = options.get("sheet_name", "报表数据")
        header_style = options.get("header_style", {})
        column_width = options.get("column_width", 18)
        freeze_panes = options.get("freeze_panes")
        report_data = report.report_data
        if not report_data:
            buffer = io.BytesIO()
            wb.save(buffer)
            return buffer.getvalue()
        columns = report_data.get("columns", [])
        rows = report_data.get("rows", [])
        summary = report_data.get("summary", {})
        header_font_config = header_style
        header_font = Font(
            bold=header_font_config.get("bold", True),
            size=header_font_config.get("size", 12)
        )
        header_fill = PatternFill(
            start_color=header_font_config.get("bg_color", "4472C4"),
            end_color=header_font_config.get("bg_color", "4472C4"),
            fill_type="solid"
        )
        header_font_white = Font(
            bold=header_font_config.get("bold", True),
            size=header_font_config.get("size", 12),
            color=header_font_config.get("color", "FFFFFF")
        )
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        ws.append(columns)
        for col_num, _ in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        for row_idx, row_data in enumerate(rows, start=2):
            row_values = [row_data.get(col, "") for col in columns]
            ws.append(row_values)
            for col_num, value in enumerate(row_values, 1):
                cell = ws.cell(row=row_idx, column=col_num)
                cell.border = thin_border
        if summary:
            ws.append([])
            summary_row = len(rows) + 3
            ws.cell(row=summary_row, column=1, value="统计汇总")
            for i, (key, value) in enumerate(summary.items(), start=1):
                ws.cell(row=summary_row, column=i + 1, value=f"{key}: {value}")
        for col_num in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = column_width
        if freeze_panes:
            ws.freeze_panes = freeze_panes
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _export_csv(self, report: Report, options: Dict[str, Any]) -> bytes:
        buffer = io.StringIO()
        delimiter = options.get("delimiter", ",")
        encoding = options.get("encoding", "utf-8")
        include_header = options.get("include_header", True)
        quote_char = options.get("quote_char", '"')
        writer = csv.writer(buffer, delimiter=delimiter, quotechar=quote_char, quoting=csv.QUOTE_MINIMAL)
        report_data = report.report_data
        if not report_data:
            return buffer.getvalue().encode(encoding)
        columns = report_data.get("columns", [])
        rows = report_data.get("rows", [])
        if include_header:
            writer.writerow(columns)
        for row_data in rows:
            writer.writerow([row_data.get(col, "") for col in columns])
        return buffer.getvalue().encode(encoding)

    def _export_pdf(self, report: Report, options: Dict[str, Any]) -> bytes:
        from reportlab.lib.pagesizes import A4, letter
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        page_size_map = {"A4": A4, "letter": letter}
        page_size = page_size_map.get(options.get("page_size", "A4"), A4)
        if options.get("orientation") == "landscape":
            page_size = (page_size[1], page_size[0])
        header_color = options.get("header_color", "#4472C4")
        alternate_row_color = options.get("alternate_row_color", "#F2F2F2")
        title_style_config = options.get("title_style", {})
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=page_size)
        styles = getSampleStyleSheet()
        story = []
        title = Paragraph(f"<b>{report.report_name}</b>", styles["Title"])
        story.append(title)
        story.append(Spacer(1, 12))
        report_data = report.report_data
        if not report_data:
            doc.build(story)
            return buffer.getvalue()
        columns = report_data.get("columns", [])
        rows = report_data.get("rows", [])
        summary = report_data.get("summary", {})
        table_data = [columns]
        for row_data in rows:
            table_data.append([str(row_data.get(col, "")) for col in columns])
        if summary:
            table_data.append([])
            table_data.append(["统计汇总"] + [f"{k}: {v}" for k, v in list(summary.items())[:len(columns) - 1]])
        table = Table(table_data)
        table_style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(header_color)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), title_style_config.get("font", "Helvetica-Bold")),
            ("FONTSIZE", (0, 0), (-1, 0), title_style_config.get("size", 10)),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
        ])
        if alternate_row_color:
            for i in range(1, len(table_data)):
                if i % 2 == 0:
                    table_style.add("BACKGROUND", (0, i), (-1, i), colors.HexColor(alternate_row_color))
        table.setStyle(table_style)
        story.append(table)
        doc.build(story)
        return buffer.getvalue()

    def register_format_handler(self, format_name: str, handler: Callable) -> None:
        self._format_handlers[format_name.lower()] = handler

    def create_export_config(self, template_id: str, export_formats: list,
                             export_options: Optional[Dict[str, Any]] = None) -> ExportConfig:
        export_id = f"export_{uuid.uuid4().hex[:12]}"
        config = ExportConfig(
            export_id=export_id,
            template_id=template_id,
            export_formats=export_formats,
            export_options=export_options or {}
        )
        self.db.add(config)
        self.db.commit()
        self.db.refresh(config)
        return config

    def get_export_config(self, template_id: str) -> Optional[ExportConfig]:
        return self.db.query(ExportConfig).filter(ExportConfig.template_id == template_id).first()


class ConfiguredRetryExportModule:
    def __init__(self, db: Session, storage_module: StorageModule,
                 statistics_module: StatisticsModule):
        self.db = db
        self.storage_module = storage_module
        self.statistics_module = statistics_module
        self.export_module = ConfigurableExportModule(db, storage_module, statistics_module)
        self._export_attempts: Dict[str, int] = {}
        self._failed_notifications: List[Dict[str, Any]] = []

    def _calculate_retry_delay(self, attempt: int, report_complexity: int = 1) -> float:
        retry_config = settings.get_retry_config(report_complexity)
        base_delay = retry_config.base_delay * (retry_config.backoff_multiplier ** attempt)
        return base_delay

    def _get_report_complexity(self, report: Report) -> int:
        if not report.report_data:
            return 1
        rows = report.report_data.get("rows", [])
        row_count = len(rows)
        if row_count < 100:
            return 1
        elif row_count < 1000:
            return 2
        elif row_count < 10000:
            return 3
        elif row_count < 100000:
            return 4
        else:
            return 5

    def _get_retry_config_for_report(self, report: Report) -> Dict[str, Any]:
        complexity = self._get_report_complexity(report)
        config = settings.get_retry_config(complexity)
        return {
            "complexity": complexity,
            "base_delay": config.base_delay,
            "max_retries": config.max_retries,
            "backoff_multiplier": config.backoff_multiplier,
            "description": config.description
        }

    def _send_failure_notification(self, report_id: str, error_message: str,
                                   attempt_count: int, complexity: int) -> None:
        notification = {
            "report_id": report_id,
            "error_message": error_message,
            "attempt_count": attempt_count,
            "complexity_level": complexity,
            "notified_at": datetime.utcnow().isoformat()
        }
        self._failed_notifications.append(notification)

    def export_with_retry(self, report: Report, export_format: str,
                          export_options: Optional[Dict[str, Any]] = None,
                          custom_max_retries: Optional[int] = None) -> str:
        retry_info = self._get_retry_config_for_report(report)
        max_retries = custom_max_retries or retry_info["max_retries"]
        complexity = retry_info["complexity"]
        last_error = None
        for attempt in range(max_retries + 1):
            try:
                file_path = self.export_module.export_report(
                    report, export_format, export_options
                )
                self._export_attempts[report.report_id] = attempt
                return file_path
            except Exception as e:
                last_error = str(e)
                if attempt < max_retries:
                    delay = self._calculate_retry_delay(attempt, complexity)
                    time.sleep(delay)
                else:
                    self._send_failure_notification(
                        report.report_id, last_error, max_retries, complexity
                    )
                    raise RuntimeError(
                        f"Export failed after {max_retries} retries "
                        f"(complexity: {complexity}). Last error: {last_error}"
                    )

    def get_export_attempts(self, report_id: str) -> Optional[int]:
        return self._export_attempts.get(report_id)

    def get_failed_notifications(self) -> List[Dict[str, Any]]:
        return self._failed_notifications.copy()

    def clear_failed_notifications(self) -> None:
        self._failed_notifications.clear()


class AsyncConfigurableExportModule:
    def __init__(self, db: Session, storage_module: StorageModule,
                 statistics_module: StatisticsModule,
                 task_queue: Optional[TaskQueue] = None,
                 use_redis: bool = None):
        self.db = db
        self.storage_module = storage_module
        self.statistics_module = statistics_module
        if use_redis is None:
            use_redis = is_redis_available()
        self.use_redis = use_redis
        if use_redis:
            self.redis_queue = RedisExportQueue()
            self.task_queue = None
        else:
            self.redis_queue = None
            self.task_queue = task_queue or TaskQueue(max_workers=5)
        self.retry_module = ConfiguredRetryExportModule(db, storage_module, statistics_module)
        self._register_handlers()

    def _register_handlers(self):
        if self.use_redis:
            return
        async def export_handler(payload):
            from reporthub.modules.query_module import QueryModule
            query_module = QueryModule(self.db)
            report_id = payload["report_id"]
            export_format = payload["export_format"]
            export_options = payload.get("export_options")
            max_retries = payload.get("max_retries")
            report = query_module.get_report_by_id(report_id)
            if not report:
                raise Exception(f"Report not found: {report_id}")
            file_path = self.retry_module.export_with_retry(
                report, export_format, export_options, max_retries
            )
            return {
                "report_id": report_id,
                "export_file": file_path,
                "export_format": export_format,
                "attempt_count": self.retry_module.get_export_attempts(report_id) or 0
            }
        self.task_queue.register_handler("export_report", export_handler)

    async def export_report_async(self, report_id: str, export_format: str,
                                   export_options: Optional[Dict[str, Any]] = None,
                                   max_retries: Optional[int] = None,
                                   report_complexity: Optional[int] = None,
                                   priority: bool = False) -> str:
        from reporthub.modules.query_module import QueryModule
        query_module = QueryModule(self.db)
        report = query_module.get_report_by_id(report_id)
        if report_complexity is None and report:
            report_complexity = self.retry_module._get_report_complexity(report)
        complexity = report_complexity or 1
        if self.use_redis:
            return self.redis_queue.submit_task(
                report_id=report_id,
                export_format=export_format,
                export_options=export_options,
                max_retries=max_retries,
                report_complexity=complexity,
                priority=priority
            )
        payload = {
            "report_id": report_id,
            "export_format": export_format,
            "export_options": export_options,
            "max_retries": max_retries
        }
        return await self.task_queue.submit_task(
            task_type="export_report",
            payload=payload,
            max_retries=max_retries or settings.get_retry_config(complexity).max_retries
        )

    def get_task_status(self, task_id: str):
        if self.use_redis:
            task_data = self.redis_queue.get_task_status(task_id)
            if not task_data:
                return None
            class RedisTaskStatus:
                def __init__(self, data: Dict[str, Any]):
                    self.task_id = data.get("task_id")
                    self.status = TaskStatus(data.get("status", "pending"))
                    self.progress = 0.0
                    self.result = data.get("export_file") or data.get("result")
                    self.error = data.get("error")
                    self.retry_count = data.get("retry_count", 0)
                    self.duration_ms = 0.0
                    self.message = data.get("message", "")
                    self.created_at = data.get("created_at")
                    self.started_at = data.get("started_at")
                    self.completed_at = data.get("completed_at")
            return RedisTaskStatus(task_data)
        return self.task_queue.get_task_status(task_id)

    def get_task_progress(self, task_id: str) -> Dict[str, Any]:
        if self.use_redis:
            task = self.get_task_status(task_id)
            if not task:
                return {"exists": False}
            return {
                "exists": True,
                "task_id": task.task_id,
                "status": task.status.value,
                "progress": task.progress,
                "message": task.message,
                "retry_count": task.retry_count,
                "created_at": task.created_at,
                "started_at": task.started_at,
                "completed_at": task.completed_at,
                "duration_ms": task.duration_ms
            }
        return self.task_queue.get_task_progress(task_id)

    async def wait_for_task(self, task_id: str, timeout: int = 300) -> Optional[Dict[str, Any]]:
        import asyncio
        start_time = time.time()
        while time.time() - start_time < timeout:
            status = self.get_task_status(task_id)
            if status and status.status in [TaskStatus.COMPLETED, TaskStatus.FAILED]:
                return {
                    "status": status.status.value,
                    "result": status.result,
                    "error": status.error,
                    "retry_count": status.retry_count,
                    "duration_ms": status.duration_ms
                }
            await asyncio.sleep(0.5)
        raise TimeoutError(f"Task {task_id} timed out after {timeout} seconds")


class ReportExportWorker:
    def __init__(self, db: Session, storage_module: StorageModule, statistics_module: StatisticsModule):
        self.db = db
        self.storage_module = storage_module
        self.statistics_module = statistics_module
        self.redis_queue = RedisExportQueue()
        self.retry_module = ConfiguredRetryExportModule(db, storage_module, statistics_module)
        self.running = False

    def start(self, max_tasks: Optional[int] = None):
        self.running = True
        processed = 0
        while self.running:
            if max_tasks and processed >= max_tasks:
                break
            task = self.redis_queue.get_next_task(timeout=1)
            if not task:
                continue
            try:
                result = self._process_task(task)
                self.redis_queue.complete_task(task["task_id"], result.get("export_file", ""))
                processed += 1
            except Exception as e:
                self.redis_queue.fail_task(task["task_id"], str(e))
                processed += 1

    def stop(self):
        self.running = False

    def _process_task(self, task: Dict[str, Any]) -> Dict[str, Any]:
        from reporthub.modules.query_module import QueryModule
        query_module = QueryModule(self.db)
        report_id = task["report_id"]
        export_format = task["export_format"]
        export_options = task.get("export_options")
        max_retries = task.get("max_retries")
        report = query_module.get_report_by_id(report_id)
        if not report:
            raise Exception(f"Report not found: {report_id}")
        file_path = self.retry_module.export_with_retry(
            report, export_format, export_options, max_retries
        )
        return {
            "report_id": report_id,
            "export_file": file_path,
            "export_format": export_format,
            "attempt_count": self.retry_module.get_export_attempts(report_id) or 0
        }


ExportModule = ConfigurableExportModule
RetryExportModule = ConfiguredRetryExportModule
AsyncExportModule = AsyncConfigurableExportModule
