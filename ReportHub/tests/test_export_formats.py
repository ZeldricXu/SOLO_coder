import pytest
import os
import csv
import io
from typing import Dict, Any
from openpyxl import load_workbook

from reporthub.modules import ExportModule
from tests.data import TestDataBuilder


class TestExcelExport:
    def test_xlsx_export_basic(self, export_module, in_memory_db, test_builder, temp_storage_path):
        template = test_builder.create_mock_template(in_memory_db)
        report = test_builder.create_mock_report(in_memory_db, template.template_id)
        file_path = export_module.export_report(report, "xlsx")
        assert file_path is not None
        assert file_path.endswith(".xlsx")
        assert os.path.exists(file_path)
        assert os.path.getsize(file_path) > 0

    def test_xlsx_export_content(self, export_module, in_memory_db, test_builder, temp_storage_path):
        template = test_builder.create_mock_template(in_memory_db)
        report = test_builder.create_mock_report(in_memory_db, template.template_id)
        file_path = export_module.export_report(report, "xlsx")
        wb = load_workbook(file_path)
        assert wb is not None
        ws = wb.active
        assert ws.title == "报表数据"
        headers = [cell.value for cell in ws[1]]
        assert "date" in headers
        assert "sales" in headers
        assert ws.max_row > 1
        assert ws.max_column >= 5
        wb.close()

    def test_xlsx_export_custom_sheet_name(self, export_module, in_memory_db, test_builder, temp_storage_path):
        template = test_builder.create_mock_template(in_memory_db)
        report = test_builder.create_mock_report(in_memory_db, template.template_id)
        custom_sheet_name = "SalesData_2026"
        file_path = export_module.export_report(
            report,
            "xlsx",
            export_options={"xlsx": {"sheet_name": custom_sheet_name}}
        )
        wb = load_workbook(file_path)
        assert custom_sheet_name in wb.sheetnames
        wb.close()

    def test_xlsx_export_header_formatting(self, export_module, in_memory_db, test_builder, temp_storage_path):
        template = test_builder.create_mock_template(in_memory_db)
        report = test_builder.create_mock_report(in_memory_db, template.template_id)
        file_path = export_module.export_report(report, "xlsx")
        wb = load_workbook(file_path)
        ws = wb.active
        header_cell = ws.cell(row=1, column=1)
        assert header_cell.font.bold is True
        assert header_cell.font.color.rgb.endswith("FFFFFF")
        assert header_cell.fill.start_color.rgb.endswith("4472C4")
        wb.close()

    def test_xlsx_export_summary(self, export_module, in_memory_db, test_builder, temp_storage_path):
        template = test_builder.create_mock_template(in_memory_db)
        report = test_builder.create_mock_report(in_memory_db, template.template_id, row_count=5)
        file_path = export_module.export_report(report, "xlsx")
        wb = load_workbook(file_path)
        ws = wb.active
        found_summary = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            row_values = [cell.value for cell in row]
            if any("统计汇总" in str(v) if v else False for v in row_values):
                found_summary = True
                break
        assert found_summary is True
        wb.close()


class TestCsvExport:
    def test_csv_export_basic(self, export_module, in_memory_db, test_builder, temp_storage_path):
        template = test_builder.create_mock_template(in_memory_db)
        report = test_builder.create_mock_report(in_memory_db, template.template_id)
        file_path = export_module.export_report(report, "csv")
        assert file_path is not None
        assert file_path.endswith(".csv")
        assert os.path.exists(file_path)

    def test_csv_export_content(self, export_module, in_memory_db, test_builder, temp_storage_path):
        template = test_builder.create_mock_template(in_memory_db)
        report = test_builder.create_mock_report(in_memory_db, template.template_id, row_count=3)
        file_path = export_module.export_report(report, "csv")
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            rows = list(reader)
        assert len(rows) >= 4
        headers = rows[0]
        assert "date" in headers
        assert "sales" in headers
        data_rows = rows[1:4]
        assert len(data_rows) == 3

    def test_csv_export_custom_delimiter(self, export_module, in_memory_db, test_builder, temp_storage_path):
        template = test_builder.create_mock_template(in_memory_db)
        report = test_builder.create_mock_report(in_memory_db, template.template_id)
        file_path = export_module.export_report(
            report,
            "csv",
            export_options={"csv": {"delimiter": ";"}}
        )
        with open(file_path, 'r', encoding='utf-8') as f:
            first_line = f.readline()
        assert ";" in first_line
        assert "," not in first_line[:first_line.index(";") + 1] if ";" in first_line else True

    def test_csv_export_encoding(self, export_module, in_memory_db, test_builder, temp_storage_path):
        template = test_builder.create_mock_template(in_memory_db)
        report = test_builder.create_mock_report(in_memory_db, template.template_id)
        file_path = export_module.export_report(
            report,
            "csv",
            export_options={"csv": {"encoding": "utf-8"}}
        )
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        assert len(content) > 0


class TestPdfExport:
    def test_pdf_export_basic(self, export_module, in_memory_db, test_builder, temp_storage_path):
        template = test_builder.create_mock_template(in_memory_db)
        report = test_builder.create_mock_report(in_memory_db, template.template_id)
        file_path = export_module.export_report(report, "pdf")
        assert file_path is not None
        assert file_path.endswith(".pdf")
        assert os.path.exists(file_path)
        assert os.path.getsize(file_path) > 0

    def test_pdf_export_signature(self, export_module, in_memory_db, test_builder, temp_storage_path):
        template = test_builder.create_mock_template(in_memory_db)
        report = test_builder.create_mock_report(in_memory_db, template.template_id)
        file_path = export_module.export_report(report, "pdf")
        with open(file_path, 'rb') as f:
            header = f.read(5)
        assert header == b'%PDF-'

    def test_pdf_export_page_size_a4(self, export_module, in_memory_db, test_builder, temp_storage_path):
        template = test_builder.create_mock_template(in_memory_db)
        report = test_builder.create_mock_report(in_memory_db, template.template_id)
        file_path = export_module.export_report(
            report,
            "pdf",
            export_options={"pdf": {"page_size": "A4", "orientation": "portrait"}}
        )
        assert os.path.exists(file_path)
        assert os.path.getsize(file_path) > 0

    def test_pdf_export_landscape(self, export_module, in_memory_db, test_builder, temp_storage_path):
        template = test_builder.create_mock_template(in_memory_db)
        report = test_builder.create_mock_report(in_memory_db, template.template_id)
        file_path = export_module.export_report(
            report,
            "pdf",
            export_options={"pdf": {"page_size": "A4", "orientation": "landscape"}}
        )
        assert os.path.exists(file_path)
        assert os.path.getsize(file_path) > 0

    def test_pdf_export_letter_size(self, export_module, in_memory_db, test_builder, temp_storage_path):
        template = test_builder.create_mock_template(in_memory_db)
        report = test_builder.create_mock_report(in_memory_db, template.template_id)
        file_path = export_module.export_report(
            report,
            "pdf",
            export_options={"pdf": {"page_size": "letter"}}
        )
        assert os.path.exists(file_path)
        assert os.path.getsize(file_path) > 0


class TestExportFormatOptions:
    def test_all_formats_exportable(self, export_module, in_memory_db, test_builder, temp_storage_path):
        template = test_builder.create_mock_template(in_memory_db)
        report = test_builder.create_mock_report(in_memory_db, template.template_id)
        formats = ["xlsx", "csv", "pdf"]
        exported_files = []
        for fmt in formats:
            file_path = export_module.export_report(report, fmt)
            assert os.path.exists(file_path)
            exported_files.append(file_path)
        assert len(exported_files) == 3

    def test_export_config_loading(self, export_module, in_memory_db, test_builder):
        template = test_builder.create_mock_template(in_memory_db)
        config = export_module.create_export_config(
            template_id=template.template_id,
            export_formats=["xlsx", "pdf", "csv"],
            export_options={
                "xlsx": {
                    "sheet_name": "报表数据",
                    "header_style": "bold"
                },
                "pdf": {
                    "page_size": "A4",
                    "orientation": "landscape"
                },
                "csv": {
                    "delimiter": ",",
                    "encoding": "utf-8"
                }
            }
        )
        assert config is not None
        assert "xlsx" in config.export_options
        assert "pdf" in config.export_options
        assert "csv" in config.export_options
        assert config.export_options["pdf"]["orientation"] == "landscape"

    def test_custom_export_parameters(self, export_module, in_memory_db, test_builder, temp_storage_path):
        template = test_builder.create_mock_template(in_memory_db)
        report = test_builder.create_mock_report(in_memory_db, template.template_id)
        custom_options = {
            "xlsx": {
                "sheet_name": "Q4_Sales",
                "column_width": 25
            },
            "pdf": {
                "page_size": "A4",
                "orientation": "landscape"
            },
            "csv": {
                "delimiter": "\t",
                "encoding": "utf-8"
            }
        }
        for fmt in ["xlsx", "csv", "pdf"]:
            file_path = export_module.export_report(
                report,
                fmt,
                export_options=custom_options
            )
            assert os.path.exists(file_path)
            assert os.path.getsize(file_path) > 0


class TestLargeDataExport:
    def test_large_dataset_excel_export(self, export_module, in_memory_db, test_builder, temp_storage_path):
        template = test_builder.create_mock_template(in_memory_db)
        report = test_builder.create_mock_report(in_memory_db, template.template_id, row_count=100)
        file_path = export_module.export_report(report, "xlsx")
        assert os.path.exists(file_path)
        wb = load_workbook(file_path)
        ws = wb.active
        assert ws.max_row >= 100
        wb.close()

    def test_large_dataset_csv_export(self, export_module, in_memory_db, test_builder, temp_storage_path):
        template = test_builder.create_mock_template(in_memory_db)
        report = test_builder.create_mock_report(in_memory_db, template.template_id, row_count=100)
        file_path = export_module.export_report(report, "csv")
        assert os.path.exists(file_path)
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            rows = list(reader)
        assert len(rows) >= 100

    def test_large_dataset_pdf_export(self, export_module, in_memory_db, test_builder, temp_storage_path):
        template = test_builder.create_mock_template(in_memory_db)
        report = test_builder.create_mock_report(in_memory_db, template.template_id, row_count=50)
        file_path = export_module.export_report(report, "pdf")
        assert os.path.exists(file_path)
        assert os.path.getsize(file_path) > 0


class TestExportValidation:
    def test_empty_report_export(self, export_module, in_memory_db, storage_module, statistics_module):
        from reporthub.models.reports import Report
        empty_report = Report(
            report_id="test_empty",
            template_id="template_test",
            report_name="空报表测试",
            report_data={
                "columns": [],
                "rows": [],
                "summary": {"total_rows": 0}
            },
            report_format="xlsx",
            status="completed"
        )
        for fmt in ["xlsx", "csv", "pdf"]:
            file_path = export_module.export_report(empty_report, fmt)
            assert file_path is not None
            assert os.path.exists(file_path)

    def test_export_format_consistency(self, export_module, in_memory_db, test_builder, temp_storage_path):
        template = test_builder.create_mock_template(in_memory_db)
        report = test_builder.create_mock_report(in_memory_db, template.template_id, row_count=5)
        xlsx_path = export_module.export_report(report, "xlsx")
        csv_path = export_module.export_report(report, "csv")
        wb = load_workbook(xlsx_path)
        ws = wb.active
        xlsx_headers = [cell.value for cell in ws[1] if cell.value is not None]
        wb.close()
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            csv_headers = next(reader)
        assert xlsx_headers == csv_headers
